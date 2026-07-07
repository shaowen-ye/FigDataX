"""Excel export for digitized figure data (openpyxl, imported lazily).

One entry point: gather several figures' extracted CSVs into a single multi-sheet
workbook (one sheet per figure) plus a provenance sheet. CSV stays the primary,
universal output; this is the "export to Excel" convenience on top of it.

    export_figures(
        "paper_figures.xlsx",
        figures=[{"name": "Fig3", "csv": "/abs/fig3_extracted.csv",
                  "provenance": "Smith 2024 Fig.3 | M1 | RMSE x=0.21% y=0.15% | 2 rounds"}],
        source_name="Smith et al. 2024",
    )
"""

from __future__ import annotations

import datetime
import os
from typing import List

from . import __version__
from .core import InputError, _require

_SHEET_BAD = str.maketrans({c: "_" for c in "[]:*?/\\"})


def _bold(cell):
    from openpyxl.styles import Font
    cell.font = Font(bold=True)
    return cell


def _sheet_name(wb, name: str) -> str:
    """Excel-legal, unique, ≤31 chars."""
    base = (name or "Sheet").translate(_SHEET_BAD)[:31] or "Sheet"
    title, n = base, 2
    while title in wb.sheetnames:
        suffix = f"_{n}"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    return title


def _write_csv_sheet(wb, name: str, csv_path: str):
    pd = _require("pandas", "Excel export")
    if not os.path.exists(csv_path):
        raise InputError(f"CSV not found: {csv_path}")
    df = pd.read_csv(csv_path)
    ws = wb.create_sheet(title=_sheet_name(wb, name))
    for c, head in enumerate(df.columns, start=1):
        _bold(ws.cell(row=1, column=c, value=str(head)))
    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            if pd.isna(val):
                continue
            ws.cell(row=r, column=c, value=val.item() if hasattr(val, "item") else val)
    ws.freeze_panes = "A2"
    return ws


def export_figures(path: str, figures: List[dict], source_name: str = "") -> str:
    """Write one sheet per figure (from its extracted CSV) + a provenance sheet.

    ``figures`` — [{"name"?, "csv", "provenance"?}]. ``name`` defaults to the CSV stem.
    Returns ``path``.
    """
    if not figures:
        raise InputError("export_figures: no figures given.")
    openpyxl = _require("openpyxl", "Excel export")
    wb = openpyxl.Workbook()

    idx = wb.active
    idx.title = "Index"
    _bold(idx.cell(row=1, column=1, value="FigDataX — extracted figures"))
    idx.cell(row=2, column=1, value=f"Source: {source_name}")
    row = 4
    for col, head in enumerate(["Figure", "Sheet", "Provenance"], start=1):
        _bold(idx.cell(row=row, column=col, value=head))
    row += 1

    prov_lines = []
    for fig in figures:
        name = fig.get("name") or os.path.splitext(os.path.basename(fig["csv"]))[0]
        ws = _write_csv_sheet(wb, name, fig["csv"])
        idx.cell(row=row, column=1, value=name)
        idx.cell(row=row, column=2, value=ws.title)
        idx.cell(row=row, column=3, value=fig.get("provenance", ""))
        prov_lines.append((name, fig.get("provenance", ""), os.path.abspath(fig["csv"])))
        row += 1

    prov = wb.create_sheet(title="Provenance")
    _bold(prov.cell(row=1, column=1, value="FigDataX provenance"))
    prov.cell(row=2, column=1, value="Source")
    prov.cell(row=2, column=2, value=source_name)
    prov.cell(row=3, column=1, value="Engine version")
    prov.cell(row=3, column=2, value=__version__)
    prov.cell(row=4, column=1, value="Exported")
    prov.cell(row=4, column=2, value=datetime.datetime.now().isoformat(timespec="seconds"))
    r = 6
    for col, head in enumerate(["Figure", "Method / RMSE / rounds", "Data file"], start=1):
        _bold(prov.cell(row=r, column=col, value=head))
    for item, p, f in prov_lines:
        r += 1
        prov.cell(row=r, column=1, value=item)
        prov.cell(row=r, column=2, value=p)
        prov.cell(row=r, column=3, value=f)

    wb.save(path)
    return path
