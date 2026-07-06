"""Excel export (openpyxl).

Two entry points:
- ``export_session`` — a single digitized figure: summary + one sheet per series.
- ``export_workbook`` — the whole document: digitized figures, PDF tables, and the
  data-mentions list, each in its own sheet(s). This is what the PDF pipeline uses.
"""

from __future__ import annotations

from typing import List, Optional

from .models import ExtractionSession


def _bold(cell):
    from openpyxl.styles import Font
    cell.font = Font(bold=True)
    return cell


def _write_series_sheets(wb, session: ExtractionSession, prefix: str = ""):
    for idx, s in enumerate(session.series, start=1):
        title = (prefix + (s.name or f"Series{idx}"))[:31]
        ws = wb.create_sheet(title=title)
        for col, head in enumerate(["X", "Y", "px", "py", "confidence", "manual"], 1):
            _bold(ws.cell(row=1, column=col, value=head))
        for r, p in enumerate(s.points, start=2):
            ws.cell(row=r, column=1, value=round(p.data_x, 6))
            ws.cell(row=r, column=2, value=round(p.data_y, 6))
            ws.cell(row=r, column=3, value=round(p.px, 2))
            ws.cell(row=r, column=4, value=round(p.py, 2))
            ws.cell(row=r, column=5, value=round(p.confidence, 3))
            ws.cell(row=r, column=6, value="yes" if p.manual else "")
        ws.freeze_panes = "A2"


def export_session(session: ExtractionSession, path: str) -> str:
    """Write a summary sheet + one data sheet per series to an .xlsx file."""
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    for col, head in enumerate(["Series", "Points", "Color HSV", "Source"], 1):
        _bold(summary.cell(row=1, column=col, value=head))
    for i, s in enumerate(session.series, start=2):
        summary.cell(row=i, column=1, value=s.name)
        summary.cell(row=i, column=2, value=len(s.points))
        summary.cell(row=i, column=3, value=str(s.color_hsv))
        summary.cell(row=i, column=4, value=session.source_label or session.image_path or "")

    _write_series_sheets(wb, session)
    wb.save(path)
    return path


def export_chart_result(result, path: str, source: str = "") -> str:
    """Write a single ChartResult (box/pie/heatmap) to a one-sheet workbook."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = result.kind[:31]
    for c, head in enumerate(result.columns, start=1):
        _bold(ws.cell(row=1, column=c, value=head))
    for r, row in enumerate(result.rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    if source:
        ws.cell(row=len(result.rows) + 3, column=1, value=f"Source: {source}")
    wb.save(path)
    return path


def export_workbook(path: str,
                    sessions: Optional[List[ExtractionSession]] = None,
                    tables: Optional[List] = None,
                    mentions: Optional[List] = None,
                    source_name: str = "") -> str:
    """Write a whole-document workbook.

    ``sessions`` — digitized figures (each becomes its own series sheets, prefixed).
    ``tables``   — TableRef objects from the PDF (each a sheet).
    ``mentions`` — Mention objects (one "Data mentions" sheet).
    """
    from openpyxl import Workbook

    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    _bold(idx.cell(row=1, column=1, value="FigDataX export"))
    idx.cell(row=2, column=1, value=f"Source: {source_name}")
    row = 4
    _bold(idx.cell(row=row, column=1, value="Contents"))
    row += 1

    if sessions:
        for si, s in enumerate(sessions, start=1):
            idx.cell(row=row, column=1,
                     value=f"Figure {si}: {s.source_label or s.image_path or ''} "
                           f"({s.total_points()} points)")
            row += 1
            _write_series_sheets(wb, s, prefix=f"F{si}_")

    if tables:
        for ti, t in enumerate(tables, start=1):
            ws = wb.create_sheet(title=f"Table_p{t.page_index + 1}_{ti}"[:31])
            for r, rowvals in enumerate(t.rows, start=1):
                for c, val in enumerate(rowvals, start=1):
                    ws.cell(row=r, column=c, value=val)
            if t.rows:
                ws.freeze_panes = "A2"
            idx.cell(row=row, column=1,
                     value=f"Table (page {t.page_index + 1}): {t.shape[0]}×{t.shape[1]}")
            row += 1

    if mentions:
        ws = wb.create_sheet(title="Data mentions")
        for col, head in enumerate(["Page", "Category", "Match", "Context"], 1):
            _bold(ws.cell(row=1, column=col, value=head))
        for r, m in enumerate(mentions, start=2):
            ws.cell(row=r, column=1, value=m.page_label)
            ws.cell(row=r, column=2, value=m.category)
            ws.cell(row=r, column=3, value=m.match_text)
            ws.cell(row=r, column=4, value=m.sentence)
        ws.freeze_panes = "A2"
        ws.column_dimensions["D"].width = 80
        idx.cell(row=row, column=1, value=f"Data mentions: {len(mentions)} flagged")
        row += 1

    wb.save(path)
    return path
