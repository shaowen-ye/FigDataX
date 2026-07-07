"""Figure-workbook export: build from CSVs, reload with openpyxl, assert."""

import os

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts.figdatax import export_figures
from scripts.figdatax import InputError


def _csv(path, df):
    df.to_csv(path, index=False)
    return str(path)


def test_workbook_sheets_and_provenance(tmp_path):
    fig_csv = _csv(tmp_path / "fig3.csv",
                   pd.DataFrame({"X": [0, 1, 2], "Y": [0.0, 2.0, 4.0]}))
    out = str(tmp_path / "figs.xlsx")

    export_figures(out,
                   figures=[{"name": "Fig3", "csv": fig_csv,
                             "provenance": "paper.pdf p.5 | M1 | RMSE x=0.2% | 2 rounds"}],
                   source_name="paper.pdf")

    assert os.path.getsize(out) > 0
    wb = load_workbook(out)
    assert "Index" in wb.sheetnames
    assert "Fig3" in wb.sheetnames
    assert "Provenance" in wb.sheetnames

    fig = wb["Fig3"]
    assert [c.value for c in fig[1]] == ["X", "Y"]
    assert fig["B4"].value == 4.0

    prov_text = "\n".join(str(c.value) for row in wb["Provenance"].iter_rows()
                          for c in row if c.value)
    assert "Engine version" in prov_text
    assert "RMSE x=0.2%" in prov_text


def test_missing_csv_raises(tmp_path):
    with pytest.raises(InputError):
        export_figures(str(tmp_path / "x.xlsx"),
                       figures=[{"name": "gone", "csv": str(tmp_path / "nope.csv")}])


def test_no_figures_raises(tmp_path):
    with pytest.raises(InputError):
        export_figures(str(tmp_path / "x.xlsx"), figures=[])


def test_duplicate_sheet_names_disambiguated(tmp_path):
    csv = _csv(tmp_path / "a.csv", pd.DataFrame({"X": [1]}))
    out = str(tmp_path / "dup.xlsx")
    export_figures(out, figures=[{"name": "Series", "csv": csv},
                                 {"name": "Series", "csv": csv}])
    wb = load_workbook(out)
    assert "Series" in wb.sheetnames and "Series_2" in wb.sheetnames
